"""
P&L Service — Parse Flipkart P&L xlsx, match SKUs, store report.

Design principles:
- Dynamic column parsing (no hardcoded indices — Flipkart can change format)
- Snapshot Casper expected BS/profit% at upload time (financial accuracy)
- Store unmatched SKUs with sku_pricing_id=null (no silent data loss)
- Duplicate detection by platform + period
"""

from datetime import datetime, date
from typing import Optional, Tuple
import re
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.models.pnl import PnlReport, PnlSkuRow
from app.models.sku import SkuPlatformConfig, SkuPricing
from app.models.platform import Platform
from app.schemas.pnl import PnlUploadResult, PnlDuplicateInfo
from app.services.fraud import (
    extract_order_events_fk,
    extract_order_events_meesho,
    extract_order_events_snapdeal_cpr,
    store_order_events,
    compute_sku_risk_scores,
    compute_return_reason_clusters,
    compute_state_risk_profiles,
    compute_actor_risk_profiles,
    generate_fraud_alerts,
)


# ── Column header keywords for dynamic mapping ────────────────────────────────
# Each key maps to a list of substrings to search in header text (case-insensitive)
# First match wins. This handles minor Flipkart format variations.

SKU_COL_MAP = {
    "sku_id":                   ["sku id", "sku name"],
    "gross_units":              ["gross units"],
    "rto_units":                ["rto (logistics"],          # "RTO (Logistics Return)" — avoids matching "Returned & Cancelled"
    "rvp_units":                ["rvp (customer"],           # "RVP (Customer Return)"
    "cancelled_units":          ["cancellations"],           # Row 1: "Cancellations" (col 6) — not "Returned & Cancelled Units" (col 3)
    "net_units":                ["net units"],
    "accounted_net_sales":      ["accounted net sales"],
    "commission_fee":           ["commission fee"],
    "collection_fee":           ["collection fee"],
    "fixed_fee":                ["fixed fee"],               # Fixed Fee (col 15) — was missing
    "reverse_shipping_fee":     ["reverse shipping"],
    "taxes_gst":                ["taxes (gst)", "tax gst"],
    "taxes_tcs":                ["taxes (tcs)", "tax tcs"],
    "taxes_tds":                ["taxes (tds)", "tax tds"],
    "rewards_benefits":         ["rewards & other benefits", "rewards and other benefits"],
    # Flipkart uses square brackets: "Bank Settlement [Projected]"
    "bank_settlement_projected":["bank settlement [projected]", "bank settlement projected"],
    "input_tax_credits":        ["input tax credits (inr)", "input tax credit"],
    "net_earnings":             ["net earnings"],
    "earnings_per_unit":        ["earnings per unit"],
    "net_margin_pct":           ["net margins", "net margin"],
    "amount_settled":           ["amount settled"],
    "amount_pending":           ["amount pending"],
}

# Summary sheet row keywords → field names
SUMMARY_ROW_MAP = {
    "gross_sales":      ["gross sales"],
    "gross_units":      ["gross sales"],           # units in col C
    "returns_amount":   ["returns & cancellations", "returns and cancellations"],
    "returned_units":   ["returns & cancellations", "returns and cancellations"],
    "net_sales":        ["accounted net sales", "estimated net sales"],
    "net_units":        ["accounted net sales", "estimated net sales"],
    "total_expenses":   ["total expenses"],
    "bank_settlement":  ["bank settlement (projected)", "bank settlement projected"],
    "input_tax_credits":["input tax credit"],
    "net_earnings":     ["earnings on platform", "net earnings"],
    "net_margin_pct":   ["net margin"],
    "amount_settled":   ["already paid"],
    "amount_pending":   ["pending"],
}


def _safe_float(val) -> Optional[float]:
    """Convert cell value to float, handling None, strings with commas, ₹, % signs, dashes."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("₹", "").replace("%", "").strip()
    if s in ("", "-", "N/A", "NA"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _safe_int(val) -> Optional[int]:
    f = _safe_float(val)
    return int(f) if f is not None else None


def _safe_pct(val) -> Optional[float]:
    """
    Convert a percentage value to 0-100 scale.
    - String '80.71%' → 80.71
    - Excel decimal 0.8071 (openpyxl returns this for % formatted cells) → 80.71
    - Plain number 80.71 (Flipkart SKU sheet stores it this way) → 80.71
    Threshold: if abs(val) <= 1.0 treat as decimal; otherwise already 0-100 scale.
    """
    if val is None:
        return None
    if isinstance(val, str) and "%" in val:
        f = _safe_float(val)
        return round(f, 2) if f is not None else None
    f = _safe_float(val)
    if f is None:
        return None
    # Only multiply if it looks like a decimal fraction (e.g. 0.8071)
    if abs(f) <= 1.0:
        return round(f * 100, 2)
    return round(f, 2)


def _parse_date_str(s: str) -> Optional[date]:
    """Try multiple date formats Flipkart uses."""
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d %b %Y", "%d/%m/%Y", "%d-%m-%Y", "%b %d, %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_period_from_summary(ws) -> Tuple[Optional[date], Optional[date]]:
    """
    Extract period_start and period_end from Sheet 1.
    Flipkart label: "Orders Recieved During:" (note: typo in Flipkart's report)
    Also handles: "Date Range", "Period", etc.
    Value format: "2026-03-01 to 2026-03-31" in col B.

    IMPORTANT: must specify max_col in iter_rows — openpyxl read_only mode
    only returns cells that have data in XML; without max_col sparse rows
    return only col A.
    """
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=4, values_only=True):
        label = str(row[0] or "").lower().strip()
        # Match Flipkart's "Orders Recieved During:" and generic "Date Range"/"Period"
        if not any(k in label for k in ["during", "date range", "period start", "report period"]):
            continue

        # Col B might be a date object (openpyxl auto-detects) or a string
        val_b = row[1] if len(row) > 1 else None
        val_c = row[2] if len(row) > 2 else None

        # Case 1: openpyxl returned actual date objects
        if isinstance(val_b, date) and isinstance(val_c, date):
            return val_b, val_c

        # Case 2: string range "2026-03-01 to 2026-03-31" or "01 Mar 2026 - 31 Mar 2026"
        val_str = str(val_b or "").strip()
        for sep in [" to ", " - ", " – ", " ~ ", ","]:
            if sep in val_str:
                parts = val_str.split(sep, 1)
                d1 = _parse_date_str(parts[0].strip())
                d2 = _parse_date_str(parts[1].strip())
                if d1 and d2:
                    return d1, d2

        # Case 3: start date in col B, end date in col C (separate cells)
        if val_b and val_c:
            d1 = _parse_date_str(str(val_b).strip())
            d2 = _parse_date_str(str(val_c).strip())
            if d1 and d2:
                return d1, d2

    return None, None


def extract_period_from_bytes(file_bytes: bytes) -> Tuple[date, date]:
    """
    Lightweight — open workbook, read only Sheet 1 rows 1-20, extract period.
    Called before full parse to enable duplicate detection.
    Raises ValueError if period cannot be found.
    """
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        start, end = _parse_period_from_summary(ws)
    finally:
        wb.close()

    if not start or not end:
        raise ValueError(
            "Could not find the report period in Sheet 1. "
            "Make sure this is a Flipkart P&L report (Sheet 1 must have an 'Orders Recieved During:' row)."
        )
    return start, end


def _build_col_index(ws) -> dict[str, int]:
    """
    Read Sheet 2 headers (rows 1 and 2) and build field_name → col_index mapping.
    Uses iter_rows (compatible with read_only=True — ws.cell() is NOT available in read_only mode).
    Merges both header rows into one combined string per column for matching.
    """
    header_rows = []
    # max_col MUST be set — read_only mode only returns cells present in XML
    # (sparse rows appear as single-element). 60 covers all known Flipkart columns.
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=60, values_only=True):
        header_rows.append([str(v or "").lower().strip() for v in row])

    if len(header_rows) < 2:
        return {}

    row1, row2 = header_rows[0], header_rows[1]
    max_cols = max(len(row1), len(row2))

    combined = []
    for i in range(max_cols):
        r1 = row1[i] if i < len(row1) else ""
        r2 = row2[i] if i < len(row2) else ""
        combined.append(f"{r1} {r2}".strip())

    mapping = {}
    for field, keywords in SKU_COL_MAP.items():
        for col_idx, header in enumerate(combined):
            if any(kw in header for kw in keywords):
                if field not in mapping:  # first match wins
                    mapping[field] = col_idx  # 0-indexed
    return mapping


def _parse_summary_sheet(ws) -> dict:
    """
    Parse Sheet 1 (Overall Summary).
    Format: Col A = label, Col B = amount INR, Col C = units.
    Searches for keyword rows rather than fixed row numbers.
    """
    data = {}
    # max_row AND max_col MUST be set — Flipkart xlsx has broken dimension attribute
    # (ws.max_row = 1 even though there are 60+ rows). max_row=200 covers all known layouts.
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=4, values_only=True):
        label = str(row[0] or "").lower().strip()
        amount = _safe_float(row[1]) if len(row) > 1 else None
        units = _safe_int(row[2]) if len(row) > 2 else None

        if not label:
            continue

        if any(k in label for k in ["gross sales"]):
            data["gross_sales"] = amount
            data["gross_units"] = units

        elif any(k in label for k in ["returns & cancellations", "returns and cancellations"]):
            data["returns_amount"] = amount
            data["returned_units"] = abs(units) if units else None

        elif "accounted net sales" in label:
            data.setdefault("net_sales", amount)

        elif "estimated net sales" in label:
            data.setdefault("net_sales", amount)
            data.setdefault("net_units", units)

        elif "total expenses" in label:
            data["total_expenses"] = amount

        elif "bank settlement (projected)" in label or "bank settlement projected" in label:
            data["bank_settlement"] = amount

        elif "input tax credit" in label:
            data["input_tax_credits"] = amount

        elif "earnings on platform" in label:
            data["net_earnings"] = amount

        elif "net margin" in label:
            # Excel stores percentage cells as decimals: 80.71% → 0.8071
            data["net_margin_pct"] = _safe_pct(row[1])

        elif "already paid" in label:
            data["amount_settled"] = amount

        elif label.startswith("pending"):
            data["amount_pending"] = amount

    return data


def _parse_sku_sheet(ws, col_map: dict) -> list[dict]:
    """
    Parse Sheet 2 (SKU-level P&L). Data starts at row 3 (1-indexed).
    Returns list of raw dicts per SKU.
    """
    rows = []
    # max_row AND max_col MUST be set — Flipkart xlsx has broken dimension attribute (ws.max_row=1)
    # Use 10000 rows to handle any report size. Rows with no SKU name are skipped below.
    for row in ws.iter_rows(min_row=3, max_row=10000, min_col=1, max_col=60, values_only=True):
        sku_col = col_map.get("sku_id", 0)
        sku_name = str(row[sku_col] or "").strip() if len(row) > sku_col else ""
        if not sku_name or sku_name.lower() in ("sku id", "sku name", "total", ""):
            continue

        def get(field):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        rows.append({
            "platform_sku_name":        sku_name,
            "gross_units":              _safe_int(get("gross_units")),
            "rto_units":                _safe_int(get("rto_units")),
            "rvp_units":                _safe_int(get("rvp_units")),
            "cancelled_units":          _safe_int(get("cancelled_units")),
            "net_units":                _safe_int(get("net_units")),
            "accounted_net_sales":      _safe_float(get("accounted_net_sales")),
            "commission_fee":           _safe_float(get("commission_fee")),
            "collection_fee":           _safe_float(get("collection_fee")),
            "fixed_fee":                _safe_float(get("fixed_fee")),
            "reverse_shipping_fee":     _safe_float(get("reverse_shipping_fee")),
            "taxes_gst":                _safe_float(get("taxes_gst")),
            "taxes_tcs":                _safe_float(get("taxes_tcs")),
            "taxes_tds":                _safe_float(get("taxes_tds")),
            "rewards_benefits":         _safe_float(get("rewards_benefits")),
            "bank_settlement_projected":_safe_float(get("bank_settlement_projected")),
            "input_tax_credits":        _safe_float(get("input_tax_credits")),
            "net_earnings":             _safe_float(get("net_earnings")),
            "earnings_per_unit":        _safe_float(get("earnings_per_unit")),
            "net_margin_pct":           _safe_pct(get("net_margin_pct")),
            "amount_settled":           _safe_float(get("amount_settled")),
            "amount_pending":           _safe_float(get("amount_pending")),
        })
    return rows


# ── Meesho parser ─────────────────────────────────────────────────────────────

_MEESHO_DELIVERED = {"Delivered", "Shipped"}
_MEESHO_RETURN    = {"Return"}
_MEESHO_RTO       = {"RTO"}
_MEESHO_CANCELLED = {"Cancelled"}


def _parse_meesho_workbook(file_bytes: bytes) -> tuple[dict, list[dict]]:
    """
    Parse Meesho P&L xlsx using pandas.
    Returns (summary_dict, sku_rows_list) in the same shape as _parse_workbook().
    """
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    df = pd.read_excel(BytesIO(file_bytes), sheet_name="Order Payments", skiprows=[0, 2], header=0)

    # Ads Cost sheet
    try:
        ads_df = pd.read_excel(BytesIO(file_bytes), sheet_name="Ads Cost", skiprows=[0, 2], header=0)
        ads_total = float(ads_df["Total Ads Cost"].sum()) if "Total Ads Cost" in ads_df.columns else 0.0
    except Exception:
        ads_total = 0.0

    del_mask = df["Live Order Status"].isin(_MEESHO_DELIVERED)
    ret_mask = df["Live Order Status"].isin(_MEESHO_RETURN)
    rto_mask = df["Live Order Status"].isin(_MEESHO_RTO)
    can_mask = df["Live Order Status"].isin(_MEESHO_CANCELLED)
    shipped_mask = del_mask | ret_mask | rto_mask  # anything that was dispatched

    def _s(series):
        v = float(series.sum())
        return round(v, 2) if v != 0 else None

    gross_units    = int(df.loc[shipped_mask, "Quantity"].sum())
    returned_units = int(df.loc[ret_mask | rto_mask, "Quantity"].sum())
    net_units      = int(df.loc[del_mask, "Quantity"].sum())

    gross_sales     = _s(df["Total Sale Amount (Incl. Shipping & GST)"])
    returns_amount  = _s(df["Total Sale Return Amount (Incl. Shipping & GST)"])
    net_sales       = round((gross_sales or 0) + (returns_amount or 0), 2)
    bank_settlement = _s(df["Final Settlement Amount"])
    tcs_total       = _s(df["TCS"])
    tds_total       = _s(df["TDS"])
    net_earnings    = bank_settlement
    net_margin_pct  = round(net_earnings / net_sales * 100, 2) if net_sales else None

    gross_orders  = int(del_mask.sum())
    return_orders = int((ret_mask | rto_mask).sum())

    summary = {
        "gross_sales":    gross_sales,
        "gross_units":    gross_units,
        "returns_amount": returns_amount,
        "returned_units": returned_units,
        "net_sales":      net_sales,
        "net_units":      net_units,
        "bank_settlement": bank_settlement,
        "net_earnings":   net_earnings,
        "net_margin_pct": net_margin_pct,
        "tcs_amount":     tcs_total,
        "tds_amount":     tds_total,
        "marketing_fee":  round(ads_total, 2) if ads_total else None,
        "gross_orders":   gross_orders,
        "return_orders":  return_orders,
        "net_orders":     gross_orders - return_orders,
        "total_expenses": round((bank_settlement or 0) - (net_sales or 0), 2),
    }

    # ── Per-SKU rows ──────────────────────────────────────────────────────────
    sku_rows: list[dict] = []
    for sku, grp in df.groupby("Supplier SKU"):
        g_del = grp[grp["Live Order Status"].isin(_MEESHO_DELIVERED)]
        g_ret = grp[grp["Live Order Status"].isin(_MEESHO_RETURN)]
        g_rto = grp[grp["Live Order Status"].isin(_MEESHO_RTO)]
        g_can = grp[grp["Live Order Status"].isin(_MEESHO_CANCELLED)]

        g_units = int(grp[grp["Live Order Status"].isin(_MEESHO_DELIVERED | _MEESHO_RETURN | _MEESHO_RTO)]["Quantity"].sum())
        n_units = int(g_del["Quantity"].sum())

        bs  = round(float(grp["Final Settlement Amount"].sum()), 2)
        ans = round(float(g_del["Total Sale Amount (Incl. Shipping & GST)"].sum()), 2)
        tcs = round(float(grp["TCS"].sum()), 2)
        tds = round(float(grp["TDS"].sum()), 2)
        shipping     = round(float(g_del["Shipping Charge (Incl. GST)"].sum()), 2)
        ret_shipping = round(float(grp.get("Return Shipping Charge (Incl. GST)", pd.Series(dtype=float)).sum()), 2) if "Return Shipping Charge (Incl. GST)" in grp.columns else 0.0

        epu        = round(bs / n_units, 2) if n_units > 0 else None
        net_margin = round(bs / ans * 100, 2) if ans else None

        sku_rows.append({
            "platform_sku_name":         str(sku),
            "gross_units":               g_units,
            "rto_units":                 int(g_rto["Quantity"].sum()),
            "rvp_units":                 int(g_ret["Quantity"].sum()),
            "cancelled_units":           int(g_can["Quantity"].sum()),
            "net_units":                 n_units,
            "accounted_net_sales":       ans,
            "commission_fee":            0.0,
            "collection_fee":            shipping if shipping else None,
            "fixed_fee":                 None,
            "reverse_shipping_fee":      ret_shipping if ret_shipping else None,
            "taxes_gst":                 None,
            "taxes_tcs":                 tcs if tcs else None,
            "taxes_tds":                 tds if tds else None,
            "rewards_benefits":          None,
            "bank_settlement_projected": bs,
            "input_tax_credits":         None,
            "net_earnings":              bs,
            "earnings_per_unit":         epu,
            "net_margin_pct":            net_margin,
            "amount_settled":            None,
            "amount_pending":            None,
        })

    return summary, sku_rows


def _parse_snapdeal_workbook(file_bytes: bytes) -> tuple[dict, list[dict]]:
    """
    Parse Snapdeal P&L xlsx (Payment Settlement Report).
    No per-SKU data available in this report format — returns empty sku_rows.
    Returns (summary_dict, []).
    """
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    bio = BytesIO(file_bytes)

    # ── Summary sheet: seller info, opening/closing balance ──────────────────
    df_sum = pd.read_excel(bio, sheet_name="Summary", header=None)
    seller_name    = str(df_sum.iloc[6, 5]).strip() if pd.notna(df_sum.iloc[6, 5]) else None
    seller_code    = str(df_sum.iloc[7, 5]).strip() if pd.notna(df_sum.iloc[7, 5]) else None
    opening_balance = float(df_sum.iloc[11, 7]) if pd.notna(df_sum.iloc[11, 7]) else None
    closing_balance = float(df_sum.iloc[36, 7]) if pd.notna(df_sum.iloc[36, 7]) else None

    # ── Total_Suboders: gross order rows (exclude "Total" text row) ───────────
    bio.seek(0)
    df_ord = pd.read_excel(bio, sheet_name="Total_Suboders", header=0)
    df_ord = df_ord[pd.to_datetime(df_ord["Transaction Date"], errors="coerce").notna()]
    gross_sales   = round(float(df_ord["Invoice Amount"].sum()), 2)
    cod_orders    = int((df_ord["Transaction Type"] == "COD Vendor Invoice").sum())
    ncod_orders   = int((df_ord["Transaction Type"] == "NCOD Vendor Invoice").sum())
    gross_orders  = cod_orders + ncod_orders

    # ── Returns ───────────────────────────────────────────────────────────────
    bio.seek(0)
    df_ret = pd.read_excel(bio, sheet_name="Returns", header=0)
    df_ret = df_ret[pd.to_datetime(df_ret["Transaction Date"], errors="coerce").notna()]
    returns_amount = round(float(df_ret["Invoice Amount"].sum()), 2)
    return_orders  = len(df_ret)

    net_sales = round(gross_sales + returns_amount, 2)

    # ── Commission and other charges ──────────────────────────────────────────
    bio.seek(0)
    df_com = pd.read_excel(bio, sheet_name="Commission and other charges", header=0)
    commission_total    = round(float(df_com["Total Commission Amount"].sum()), 2)
    marketing_fee       = round(float(df_com["Marketing Fee"].sum()), 2) if "Marketing Fee" in df_com.columns else None
    courier_fee         = round(float(df_com["Courier Fee"].sum()), 2) if "Courier Fee" in df_com.columns else None
    payment_coll_fee    = round(float(df_com["Payment Collection Fee"].sum()), 2) if "Payment Collection Fee" in df_com.columns else None

    # ── TCS ───────────────────────────────────────────────────────────────────
    bio.seek(0)
    df_tcs = pd.read_excel(bio, sheet_name="TCS", header=0)
    tcs_amount = round(float(df_tcs["Tax Amount"].sum()), 2) if "Tax Amount" in df_tcs.columns else None

    # ── Non Order Transactions (TDS) ──────────────────────────────────────────
    bio.seek(0)
    df_non = pd.read_excel(bio, sheet_name="Non Order Transactions", header=0)
    tds_amount = round(float(df_non["Gross Amount"].sum()), 2) if "Gross Amount" in df_non.columns else None

    # ── Payments (actual bank transfers) ─────────────────────────────────────
    bio.seek(0)
    df_pay = pd.read_excel(bio, sheet_name="Payments", header=0)
    # Exclude summary "Total" row (Transaction Date = "Total" or Gross Amount = 0 at end)
    df_pay_real = df_pay[pd.to_datetime(df_pay["Transaction Date"], errors="coerce").notna()]
    bank_settlement = round(float(df_pay_real["Gross Amount"].sum()), 2)

    # ── Derived ───────────────────────────────────────────────────────────────
    tcs  = tcs_amount or 0.0
    tds  = tds_amount or 0.0
    net_earnings   = round(net_sales + commission_total + tcs + tds, 2)
    net_margin_pct = round(net_earnings / net_sales * 100, 2) if net_sales else None
    total_expenses = round(commission_total + tcs + tds, 2)
    amount_pending = round(closing_balance, 2) if closing_balance is not None else None

    summary = {
        "seller_name":          seller_name,
        "seller_code":          seller_code,
        "gross_sales":          gross_sales,
        "returns_amount":       returns_amount,
        "net_sales":            net_sales,
        "gross_orders":         gross_orders,
        "return_orders":        return_orders,
        "net_orders":           gross_orders - return_orders,
        "cod_orders":           cod_orders,
        "ncod_orders":          ncod_orders,
        "commission_total":     commission_total,
        "marketing_fee":        marketing_fee,
        "courier_fee":          courier_fee,
        "payment_collection_fee": payment_coll_fee,
        "tcs_amount":           tcs_amount,
        "tds_amount":           tds_amount,
        "bank_settlement":      bank_settlement,
        "net_earnings":         net_earnings,
        "net_margin_pct":       net_margin_pct,
        "total_expenses":       total_expenses,
        "opening_balance":      opening_balance,
        "closing_balance":      closing_balance,
        "amount_settled":       bank_settlement,
        "amount_pending":       amount_pending,
    }

    return summary, []     # no per-SKU data in this report format


def extract_period_from_bytes_snapdeal(file_bytes: bytes) -> tuple[date, date]:
    """Extract period from Snapdeal Summary sheet (row 8: Date | start | 'to' | end)."""
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    df = pd.read_excel(BytesIO(file_bytes), sheet_name="Summary", header=None)
    try:
        start_str = str(df.iloc[8, 5]).strip()   # e.g. "01-APR-2026"
        end_str   = str(df.iloc[8, 7]).strip()   # e.g. "30-APR-2026"
        period_start = pd.to_datetime(start_str, format="%d-%b-%Y").date()
        period_end   = pd.to_datetime(end_str,   format="%d-%b-%Y").date()
        return period_start, period_end
    except Exception as e:
        raise ValueError(f"Snapdeal period extraction failed: {e}")


# ── Snapdeal CPR (ComprehensivePaymentReport) parser ─────────────────────────

_CPR_DELIVERED  = {"Delivered"}
_CPR_SHIPPED    = {"Shipped", "To be Shipped"}
_CPR_RTO        = {"Courier Return"}
_CPR_RETURN     = {"Customer Return"}
_CPR_CANCELLED  = {"Seller Cancelled", "Courier Cancelled"}


def _is_snapdeal_cpr(file_bytes: bytes) -> bool:
    """Detect Snapdeal CPR format by checking for 'SubOrder Code' column in sheet 1."""
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, nrows=0)
        return "SubOrder Code" in df.columns and "SKU" in df.columns
    except Exception:
        return False


def _parse_snapdeal_cpr_workbook(file_bytes: bytes) -> tuple[dict, list[dict]]:
    """
    Parse Snapdeal ComprehensivePaymentReport (CPR) format.
    - Single flat sheet, one row per sub-order
    - Has per-SKU data: SKU, SP, Logistics, Ads, Settled, etc.
    - Commission = 0; Snapdeal monetises via Logistics + Ads Facilitation fees
    Returns (summary_dict, sku_rows_list).
    """
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    df = pd.read_excel(BytesIO(file_bytes), sheet_name=0)

    # Normalise key columns
    def _n(col):
        return pd.to_numeric(df[col], errors="coerce").fillna(0.0) if col in df.columns else pd.Series(0.0, index=df.index)

    status_col = df["Order Status"].fillna("")

    del_mask   = status_col.isin(_CPR_DELIVERED)
    ship_mask  = status_col.isin(_CPR_SHIPPED)
    rto_mask   = status_col.isin(_CPR_RTO)
    ret_mask   = status_col.isin(_CPR_RETURN)
    canc_mask  = status_col.isin(_CPR_CANCELLED)
    net_mask   = del_mask | ship_mask          # "delivered or in transit"
    return_mask = rto_mask | ret_mask
    active_mask = ~canc_mask                   # everything except cancelled

    sp           = _n("SP")
    order_amt    = _n("Order Amount")
    logistics    = _n("Logistics Fee (a)")
    ads          = _n("Ads Facilitation Fee (RO) (b)")
    net_fee      = _n("Net Charged Fee (i) = (g-h)")
    gst_fees     = _n("GST")
    tcs_col      = _n("TCS")
    tds_col      = _n("TDS")
    settled_col  = _n("Settled")

    gross_sales    = round(float(order_amt[active_mask].sum()), 2)
    returns_amount = round(-float(order_amt[return_mask].sum()), 2)  # negative
    net_sales      = round(gross_sales + returns_amount, 2)

    bank_settlement  = round(float(settled_col.sum()), 2)
    tcs_total        = round(float(tcs_col.sum()), 2)
    tds_total        = round(float(tds_col.sum()), 2)
    courier_fee_tot  = round(float(logistics.sum()), 2)
    marketing_fee_t  = round(float(ads.sum()), 2)
    total_expenses   = round(float((net_fee + gst_fees).sum()), 2)
    net_earnings     = bank_settlement
    net_margin_pct   = round(net_earnings / net_sales * 100, 2) if net_sales else None

    gross_orders  = int(active_mask.sum())
    return_orders = int(return_mask.sum())
    net_orders    = int(net_mask.sum())

    # ── Summary ───────────────────────────────────────────────────────────────
    summary = {
        "gross_sales":      gross_sales,
        "returns_amount":   returns_amount,
        "net_sales":        net_sales,
        "gross_units":      gross_orders,
        "returned_units":   return_orders,
        "net_units":        net_orders,
        "bank_settlement":  bank_settlement,
        "net_earnings":     net_earnings,
        "net_margin_pct":   net_margin_pct,
        "tcs_amount":       tcs_total or None,
        "tds_amount":       tds_total or None,
        "courier_fee":      courier_fee_tot or None,
        "marketing_fee":    marketing_fee_t or None,
        "commission_total": 0.0,
        "total_expenses":   total_expenses or None,
        "gross_orders":     gross_orders,
        "return_orders":    return_orders,
        "net_orders":       net_orders,
        "amount_settled":   bank_settlement,
    }

    # ── Per-SKU rows ──────────────────────────────────────────────────────────
    sku_rows: list[dict] = []
    for sku, grp in df.groupby("SKU"):
        if not sku or str(sku).strip() == "":
            continue

        s_col = grp["Order Status"].fillna("")

        def _cnt(mask_set):
            return int(s_col.isin(mask_set).sum())

        def _sum(col):
            return round(float(pd.to_numeric(grp[col], errors="coerce").fillna(0.0).sum()), 2) if col in grp.columns else None

        g_del    = _cnt(_CPR_DELIVERED)
        g_ship   = _cnt(_CPR_SHIPPED)
        g_rto    = _cnt(_CPR_RTO)
        g_ret    = _cnt(_CPR_RETURN)
        g_canc   = _cnt(_CPR_CANCELLED)
        g_units  = g_del + g_ship + g_rto + g_ret   # excludes cancelled
        n_units  = g_del + g_ship

        # Use Order Amount for delivered orders as revenue proxy
        net_del_mask  = s_col.isin(_CPR_DELIVERED | _CPR_SHIPPED)
        net_sales_sku = round(float(pd.to_numeric(grp.loc[net_del_mask, "Order Amount"], errors="coerce").fillna(0.0).sum()), 2)

        # Logistics split: forward (delivered/shipped) vs reverse (returns)
        ret_mask_grp   = s_col.isin(_CPR_RTO | _CPR_RETURN)
        fwd_logistics  = round(float(pd.to_numeric(grp.loc[net_del_mask, "Logistics Fee (a)"] if "Logistics Fee (a)" in grp.columns else pd.Series(dtype=float), errors="coerce").fillna(0.0).sum()), 2)
        rev_logistics  = round(float(pd.to_numeric(grp.loc[ret_mask_grp, "Logistics Fee (a)"] if "Logistics Fee (a)" in grp.columns else pd.Series(dtype=float), errors="coerce").fillna(0.0).sum()), 2)

        bs_sku  = _sum("Settled")
        tcs_sku = _sum("TCS")
        tds_sku = _sum("TDS")
        epu     = round(bs_sku / n_units, 2) if (n_units > 0 and bs_sku is not None) else None
        margin  = round(bs_sku / net_sales_sku * 100, 2) if (net_sales_sku and bs_sku is not None) else None

        sku_rows.append({
            "platform_sku_name":         str(sku).strip(),
            "gross_units":               g_units,
            "rto_units":                 g_rto,
            "rvp_units":                 g_ret,
            "cancelled_units":           g_canc,
            "net_units":                 n_units,
            "accounted_net_sales":       net_sales_sku,
            "commission_fee":            0.0,
            "collection_fee":            fwd_logistics if fwd_logistics else None,
            "fixed_fee":                 None,
            "reverse_shipping_fee":      rev_logistics if rev_logistics else None,
            "taxes_gst":                 None,
            "taxes_tcs":                 tcs_sku if tcs_sku else None,
            "taxes_tds":                 tds_sku if tds_sku else None,
            "rewards_benefits":          None,
            "bank_settlement_projected": bs_sku,
            "input_tax_credits":         None,
            "net_earnings":              bs_sku,
            "earnings_per_unit":         epu,
            "net_margin_pct":            margin,
            "amount_settled":            bs_sku,
            "amount_pending":            None,
        })

    return summary, sku_rows


def extract_period_from_bytes_snapdeal_cpr(file_bytes: bytes) -> tuple[date, date]:
    """Extract period from CPR via min/max of order_date column."""
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, usecols=["order_date"])
    dates = pd.to_datetime(df["order_date"], errors="coerce").dropna()
    if dates.empty:
        raise ValueError("CPR period extraction failed: no valid order_date values.")
    return dates.min().date(), dates.max().date()


def extract_period_from_bytes_meesho(file_bytes: bytes) -> tuple[date, date]:
    """Extract period from Meesho Excel via min/max of Order Date column."""
    from io import BytesIO
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    df = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name="Order Payments",
        skiprows=[0, 2],
        header=0,
        usecols=["Order Date"],
    )
    dates = pd.to_datetime(df["Order Date"], errors="coerce").dropna()
    if dates.empty:
        raise ValueError("Could not extract period: 'Order Date' column has no valid dates.")
    return dates.min().date(), dates.max().date()


async def check_duplicate(
    session: AsyncSession,
    platform_id: int,
    period_start,
    period_end,
    company_id: int,
) -> Optional[PnlReport]:
    """Check if a report for this company + platform + period already exists."""
    result = await session.execute(
        select(PnlReport).where(
            PnlReport.company_id == company_id,
            PnlReport.platform_id == platform_id,
            PnlReport.period_start == period_start,
            PnlReport.period_end == period_end,
        )
    )
    return result.scalar_one_or_none()


# ── Helpers for parse_and_store ───────────────────────────────────────────────

# Flipkart "actuals" — fields copied verbatim from parsed row onto PnlSkuRow
_ACTUAL_FIELDS = (
    "gross_units", "rto_units", "rvp_units", "cancelled_units", "net_units",
    "accounted_net_sales", "commission_fee", "collection_fee", "fixed_fee",
    "reverse_shipping_fee", "taxes_gst", "taxes_tcs", "taxes_tds",
    "rewards_benefits", "bank_settlement_projected", "input_tax_credits",
    "net_earnings", "earnings_per_unit", "net_margin_pct",
    "amount_settled", "amount_pending",
)


def _parse_workbook(file_bytes: bytes) -> tuple[dict, list[dict]]:
    """Load workbook and parse both sheets. Returns (summary_dict, raw_sku_rows)."""
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    summary_ws = wb[wb.sheetnames[0]]
    sku_ws     = wb[wb.sheetnames[1]]
    summary      = _parse_summary_sheet(summary_ws)
    col_map      = _build_col_index(sku_ws)
    sku_rows_raw = _parse_sku_sheet(sku_ws, col_map)
    return summary, sku_rows_raw


async def _build_pricing_lookup(session: AsyncSession, platform_id: int, company_id: int) -> dict[str, SkuPricing]:
    """
    Build a case-insensitive map: platform_sku_name.upper() → SkuPricing.
    Used to match Flipkart SKU names against this company's Casper pricing.
    """
    config_result = await session.execute(
        select(SkuPlatformConfig).where(
            SkuPlatformConfig.platform_id == platform_id,
            SkuPlatformConfig.platform_sku_name.isnot(None),
            SkuPlatformConfig.company_id == company_id,
        )
    )
    configs = config_result.scalars().all()

    pricing_ids = [c.sku_pricing_id for c in configs]
    if not pricing_ids:
        return {}

    pricing_result = await session.execute(
        select(SkuPricing).where(SkuPricing.id.in_(pricing_ids), SkuPricing.company_id == company_id)
    )
    pricing_by_id = {sp.id: sp for sp in pricing_result.scalars().all()}

    return {
        cfg.platform_sku_name.strip().upper(): pricing_by_id[cfg.sku_pricing_id]
        for cfg in configs
        if cfg.platform_sku_name and cfg.sku_pricing_id in pricing_by_id
    }


def _build_report_model(summary: dict, platform_id: int, filename: str,
                        uploaded_by: int, period_start, period_end, company_id: int) -> PnlReport:
    """Instantiate a PnlReport from parsed summary sheet data."""
    return PnlReport(
        company_id=company_id,
        platform_id=platform_id,
        period_start=period_start,
        period_end=period_end,
        filename=filename,
        uploaded_by=uploaded_by,
        uploaded_at=datetime.utcnow(),
        status="done",
        # Common fields
        gross_sales=summary.get("gross_sales"),
        gross_units=summary.get("gross_units"),
        returns_amount=summary.get("returns_amount"),
        returned_units=summary.get("returned_units"),
        net_sales=summary.get("net_sales"),
        net_units=summary.get("net_units"),
        total_expenses=summary.get("total_expenses"),
        bank_settlement=summary.get("bank_settlement"),
        input_tax_credits=summary.get("input_tax_credits"),
        net_earnings=summary.get("net_earnings"),
        net_margin_pct=summary.get("net_margin_pct"),
        amount_settled=summary.get("amount_settled"),
        amount_pending=summary.get("amount_pending"),
        # Platform-specific fields (null if not applicable)
        gross_orders=summary.get("gross_orders"),
        return_orders=summary.get("return_orders"),
        net_orders=summary.get("net_orders"),
        tcs_amount=summary.get("tcs_amount"),
        tds_amount=summary.get("tds_amount"),
        marketing_fee=summary.get("marketing_fee"),
        seller_name=summary.get("seller_name"),
        seller_code=summary.get("seller_code"),
        cod_orders=summary.get("cod_orders"),
        ncod_orders=summary.get("ncod_orders"),
        courier_fee=summary.get("courier_fee"),
        payment_collection_fee=summary.get("payment_collection_fee"),
        commission_total=summary.get("commission_total"),
        opening_balance=summary.get("opening_balance"),
        closing_balance=summary.get("closing_balance"),
    )


def _build_sku_row(raw: dict, report_id: int, matched_pricing: Optional[SkuPricing], company_id: int) -> PnlSkuRow:
    """
    Build a PnlSkuRow from parsed raw data + matched Casper pricing.
    If matched: computes variance vs Casper expected BS (snapshot at upload time).
    If unmatched: variance fields left null.
    """
    casper_fields: dict = dict(
        sku_pricing_id=None,
        casper_expected_bs=None,
        casper_expected_profit_pct=None,
        variance_bs=None,
        variance_margin_pct=None,
    )

    if matched_pricing is not None:
        sp           = matched_pricing
        actual_bs    = raw.get("bank_settlement_projected")
        net_units    = raw.get("net_units") or 0
        expected_bs  = sp.bank_settlement
        expected_tot = round(expected_bs * net_units, 2) if expected_bs else None
        variance_bs  = round(actual_bs - expected_tot, 2) if (actual_bs is not None and expected_tot is not None) else None

        actual_margin   = raw.get("net_margin_pct")
        expected_margin = sp.profit_percentage
        variance_margin = round(actual_margin - expected_margin, 2) if (actual_margin is not None and expected_margin) else None

        casper_fields = dict(
            sku_pricing_id=sp.id,
            casper_expected_bs=expected_bs,
            casper_expected_profit_pct=sp.profit_percentage,
            variance_bs=variance_bs,
            variance_margin_pct=variance_margin,
            # Freeze the cost basis at upload time so this report's profit never
            # moves when SKU pricing is edited later (closed periods stay closed).
            snap_cogs_per_unit=sp.price,
            snap_fulfillment_per_unit=(sp.package or 0) + (sp.logistics or 0) + (sp.addons or 0),
            snap_return_per_unit=(sp.cr_cost or 0) + (sp.damage_cost or 0),
            snap_overhead_per_unit=sp.misc_total,
            snap_breakeven=sp.breakeven,
            snap_gst=sp.gst,
        )

    row = PnlSkuRow(
        company_id=company_id,
        report_id=report_id,
        platform_sku_name=raw["platform_sku_name"],
        **casper_fields,
    )
    for field in _ACTUAL_FIELDS:
        setattr(row, field, raw.get(field))
    return row


async def parse_and_store(
    session: AsyncSession,
    file_bytes: bytes,
    filename: str,
    platform_id: int,
    uploaded_by: int,
    period_start,
    period_end,
    company_id: int,
    platform_name: str = "flipkart",
) -> PnlUploadResult:
    """
    Main entry point. Parse xlsx, match SKUs against Casper pricing, store report + rows.
    Dispatches to platform-specific parser based on platform_name.
    Returns upload result with match counts.
    """
    _pname = platform_name.lower()
    if _pname == "meesho":
        summary, sku_rows_raw = _parse_meesho_workbook(file_bytes)
    elif _pname == "snapdeal":
        if _is_snapdeal_cpr(file_bytes):
            summary, sku_rows_raw = _parse_snapdeal_cpr_workbook(file_bytes)
        else:
            summary, sku_rows_raw = _parse_snapdeal_workbook(file_bytes)
    elif _pname == "shopdeck":
        from app.services.shopdeck import parse_shopdeck
        summary, sku_rows_raw = parse_shopdeck(file_bytes)
    else:
        summary, sku_rows_raw = _parse_workbook(file_bytes)

    platform = await session.scalar(select(Platform).where(Platform.id == platform_id))
    platform_name = platform.name if platform else "Unknown"

    name_to_pricing = await _build_pricing_lookup(session, platform_id, company_id)

    report = _build_report_model(summary, platform_id, filename, uploaded_by, period_start, period_end, company_id)
    session.add(report)
    await session.flush()  # need report.id for child rows

    matched = unmatched = 0
    for raw in sku_rows_raw:
        sp = name_to_pricing.get(raw["platform_sku_name"].strip().upper())
        if sp is not None:
            matched += 1
        else:
            unmatched += 1
        session.add(_build_sku_row(raw, report.id, sp, company_id))

    # ── Extract + store order-level events for fraud intelligence ─────────────
    try:
        if _pname == "flipkart":
            order_events = extract_order_events_fk(file_bytes)
        elif _pname == "meesho":
            order_events = extract_order_events_meesho(file_bytes)
        elif _pname == "snapdeal" and _is_snapdeal_cpr(file_bytes):
            order_events = extract_order_events_snapdeal_cpr(file_bytes)
        else:
            order_events = []

        await store_order_events(session, order_events, report.id, platform_id, company_id)
    except Exception:
        pass  # Never fail a P&L upload due to fraud extraction error

    await session.commit()

    # ── Recompute risk scores + generate fraud alerts ─────────────────────────
    try:
        await compute_sku_risk_scores(session, platform_id, company_id)
        await session.commit()
        await compute_return_reason_clusters(session, company_id)
        await compute_state_risk_profiles(session, company_id)
        await compute_actor_risk_profiles(session, company_id)
        await generate_fraud_alerts(session, platform_id, report.id, company_id)
        await session.commit()
    except Exception:
        pass  # Risk score / alert failure is non-critical

    return PnlUploadResult(
        report_id=report.id,
        platform_name=platform_name,
        period_start=period_start,
        period_end=period_end,
        total_skus=len(sku_rows_raw),
        matched_skus=matched,
        unmatched_skus=unmatched,
        duplicate=False,
    )


async def get_all_reports(session: AsyncSession, company_id: int, platform_id: Optional[int] = None) -> list[PnlReport]:
    """List all reports for a company, optionally filtered by platform."""
    q = select(PnlReport).where(PnlReport.company_id == company_id).order_by(PnlReport.period_start.desc())
    if platform_id:
        q = q.where(PnlReport.platform_id == platform_id)
    result = await session.execute(q)
    return result.scalars().all()


async def get_report_detail(session: AsyncSession, report_id: int, company_id: int) -> Optional[PnlReport]:
    """Fetch full report, eagerly loading SKU rows + their live sku_pricing relationship."""
    from sqlalchemy.orm import selectinload
    result = await session.execute(
        select(PnlReport)
        .options(
            selectinload(PnlReport.sku_rows)
            .selectinload(PnlSkuRow.sku_pricing)
        )
        .where(PnlReport.id == report_id, PnlReport.company_id == company_id)
    )
    return result.scalar_one_or_none()


async def delete_report(session: AsyncSession, report_id: int, company_id: int) -> bool:
    """Delete report + all rows (cascade handles rows)."""
    result = await session.execute(select(PnlReport).where(PnlReport.id == report_id, PnlReport.company_id == company_id))
    report = result.scalar_one_or_none()
    if not report:
        return False
    await session.delete(report)
    await session.commit()
    return True


async def get_dashboard_summary(session: AsyncSession, company_id: int) -> dict:
    """
    Aggregate all P&L reports into dashboard-level stats.
    Returns platform totals, monthly breakdown, and overall KPIs.
    """
    result = await session.execute(
        select(PnlReport, Platform.name.label("platform_name"))
        .join(Platform, Platform.id == PnlReport.platform_id)
        .where(PnlReport.company_id == company_id)
        .order_by(PnlReport.period_start)
    )
    rows = result.all()

    if not rows:
        return {
            "platforms": [], "monthly": [],
            "total_bank_settlement": 0, "total_gross_sales": 0,
            "total_net_earnings": 0, "report_count": 0,
            "latest_period_start": None, "latest_period_end": None,
        }

    # ── Per-platform aggregation ──────────────────────────────────────────────
    plat_map: dict[int, dict] = {}
    for report, pname in rows:
        pid = report.platform_id
        if pid not in plat_map:
            plat_map[pid] = {
                "platform_id":       pid,
                "platform_name":     pname,
                "bank_settlement":   0.0,
                "gross_sales":       0.0,
                "net_sales":         0.0,
                "net_earnings":      0.0,
                "gross_units":       0,
                "net_units":         0,
                "report_count":      0,
            }
        p = plat_map[pid]
        p["bank_settlement"] += report.bank_settlement or 0
        p["gross_sales"]     += report.gross_sales or 0
        p["net_sales"]       += report.net_sales or 0
        p["net_earnings"]    += report.net_earnings or 0
        p["gross_units"]     += report.gross_units or report.gross_orders or 0
        p["net_units"]       += report.net_units or report.net_orders or 0
        p["report_count"]    += 1

    total_bs = sum(p["bank_settlement"] for p in plat_map.values())
    platforms = []
    for p in sorted(plat_map.values(), key=lambda x: -x["bank_settlement"]):
        p["bank_settlement"] = round(p["bank_settlement"], 2)
        p["gross_sales"]     = round(p["gross_sales"], 2)
        p["net_sales"]       = round(p["net_sales"], 2)
        p["net_earnings"]    = round(p["net_earnings"], 2)
        p["pct"]             = round(p["bank_settlement"] / total_bs * 100, 2) if total_bs else 0
        platforms.append(p)

    # ── Monthly breakdown ─────────────────────────────────────────────────────
    monthly_map: dict[tuple, dict] = {}
    for report, pname in rows:
        month_key = report.period_start.strftime("%Y-%m")
        key = (month_key, report.platform_id)
        if key not in monthly_map:
            monthly_map[key] = {
                "month":           month_key,
                "platform_id":     report.platform_id,
                "platform_name":   pname,
                "bank_settlement": 0.0,
                "gross_sales":     0.0,
                "report_count":    0,
            }
        m = monthly_map[key]
        m["bank_settlement"] += report.bank_settlement or 0
        m["gross_sales"]     += report.gross_sales or 0
        m["report_count"]    += 1

    monthly = [
        {**v, "bank_settlement": round(v["bank_settlement"], 2),
               "gross_sales": round(v["gross_sales"], 2)}
        for v in sorted(monthly_map.values(), key=lambda x: (x["month"], x["platform_id"]))
    ]

    all_reports = [r for r, _ in rows]
    return {
        "platforms":             platforms,
        "monthly":               monthly,
        "total_bank_settlement": round(total_bs, 2),
        "total_gross_sales":     round(sum(p["gross_sales"] for p in platforms), 2),
        "total_net_earnings":    round(sum(p["net_earnings"] for p in platforms), 2),
        "report_count":          len(all_reports),
        "latest_period_start":   str(min(r.period_start for r in all_reports)),
        "latest_period_end":     str(max(r.period_end   for r in all_reports)),
    }
