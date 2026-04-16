"""
Tests for pnl service helper functions:
_safe_float, _safe_int, _safe_pct, _build_col_index, _parse_summary_sheet
"""
import pytest
from io import BytesIO
import openpyxl
from app.services.pnl import (
    _safe_float, _safe_int, _safe_pct,
    _build_col_index, _parse_summary_sheet,
    extract_period_from_bytes,
)


# ── _safe_float ───────────────────────────────────────────────────────────────

class TestSafeFloat:
    def test_none(self):           assert _safe_float(None) is None
    def test_int(self):            assert _safe_float(100) == 100.0
    def test_float(self):          assert _safe_float(3.14) == 3.14
    def test_negative(self):       assert _safe_float(-250.5) == -250.5
    def test_string_plain(self):   assert _safe_float("324854") == 324854.0
    def test_string_comma(self):   assert _safe_float("3,24,854") == 324854.0
    def test_string_rupee(self):   assert _safe_float("₹89480.52") == 89480.52
    def test_string_pct(self):     assert _safe_float("75.76%") == 75.76
    def test_dash(self):           assert _safe_float("-") is None
    def test_empty(self):          assert _safe_float("") is None
    def test_na(self):             assert _safe_float("N/A") is None
    def test_zero(self):           assert _safe_float(0) == 0.0
    def test_zero_str(self):       assert _safe_float("0") == 0.0


# ── _safe_int ─────────────────────────────────────────────────────────────────

class TestSafeInt:
    def test_none(self):           assert _safe_int(None) is None
    def test_int(self):            assert _safe_int(813) == 813
    def test_float(self):          assert _safe_int(412.0) == 412
    def test_string(self):         assert _safe_int("412") == 412
    def test_negative(self):       assert _safe_int(-401) == -401
    def test_dash(self):           assert _safe_int("-") is None
    def test_empty(self):          assert _safe_int("") is None


# ── _safe_pct ─────────────────────────────────────────────────────────────────

class TestSafePct:
    def test_none(self):
        assert _safe_pct(None) is None

    def test_string_with_pct(self):
        # "75.76%" → 75.76 (already human-readable)
        assert _safe_pct("75.76%") == 75.76

    def test_decimal_fraction(self):
        # 0.8071 → 80.71 (Excel percentage-formatted cell)
        assert _safe_pct(0.8071) == pytest.approx(80.71, abs=0.01)

    def test_plain_number_large(self):
        # 69.88 → 69.88 (Flipkart SKU sheet stores it this way — NOT a decimal)
        assert _safe_pct(69.88) == pytest.approx(69.88, abs=0.01)

    def test_plain_number_100(self):
        # 100.0 → 100.0
        assert _safe_pct(100.0) == 100.0

    def test_negative_decimal(self):
        # -0.62 → -62.0
        assert _safe_pct(-0.62) == pytest.approx(-62.0, abs=0.01)

    def test_negative_plain(self):
        # -43.2 → -43.2 (already in 0-100 scale)
        assert _safe_pct(-43.2) == pytest.approx(-43.2, abs=0.01)

    def test_zero(self):
        assert _safe_pct(0) == 0.0

    def test_boundary_exactly_one(self):
        # 1.0 treated as decimal → 100.0
        assert _safe_pct(1.0) == pytest.approx(100.0, abs=0.01)

    def test_just_above_one(self):
        # 1.01 treated as plain → 1.01
        assert _safe_pct(1.01) == pytest.approx(1.01, abs=0.01)


# ── _build_col_index ─────────────────────────────────────────────────────────

def _make_sku_sheet(headers_row1: list, headers_row2: list, data_rows: list = None):
    """Helper: create in-memory openpyxl workbook with given headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers_row1)
    ws.append(headers_row2)
    for row in (data_rows or []):
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    return wb2.active


class TestBuildColIndex:
    def test_basic_mapping(self):
        row1 = ["SKU ID", "Gross Units (#)", "Net Units (#)", "Bank Settlement [Projected] (INR)", "Net Earnings (INR)", "Earnings per unit (INR)", "Net Margins\n(% of Net Sales value)"]
        row2 = ["", "", "", "", "", "", ""]
        ws = _make_sku_sheet(row1, row2)
        col_map = _build_col_index(ws)
        assert col_map["sku_id"] == 0
        assert col_map["gross_units"] == 1
        assert col_map["net_units"] == 2
        assert col_map["bank_settlement_projected"] == 3
        assert col_map["net_earnings"] == 4
        assert col_map["earnings_per_unit"] == 5
        assert col_map["net_margin_pct"] == 6

    def test_split_headers(self):
        # Flipkart real format: group header in row1, sub-header in row2
        row1 = ["SKU ID", None, "Returned & Cancelled Units (Breakup)", None, None, "Net Units (#)"]
        row2 = [None, None, "RTO (Logistics Return)", "RVP (Customer Return)", "Cancellations", None]
        ws = _make_sku_sheet(row1, row2)
        col_map = _build_col_index(ws)
        assert col_map["rto_units"] == 2
        assert col_map["rvp_units"] == 3
        assert col_map["cancelled_units"] == 4
        assert col_map["net_units"] == 5

    def test_fixed_fee_mapped(self):
        row1 = ["SKU ID", "Total Expenses (INR)", None, None]
        row2 = [None, "Commission Fee", "Fixed Fee", "Collection Fee"]
        ws = _make_sku_sheet(row1, row2)
        col_map = _build_col_index(ws)
        assert col_map["commission_fee"] == 1
        assert col_map["fixed_fee"] == 2
        assert col_map["collection_fee"] == 3

    def test_first_match_wins(self):
        # bank_settlement appears twice — must pick first (col 3), not second (col 5)
        row1 = ["SKU ID", None, None, "Bank Settlement [Projected] (INR)", None, "Bank Settlement [Projected] (INR)"]
        row2 = [None, None, None, None, None, None]
        ws = _make_sku_sheet(row1, row2)
        col_map = _build_col_index(ws)
        assert col_map["bank_settlement_projected"] == 3

    def test_missing_field_not_in_map(self):
        row1 = ["SKU ID"]
        row2 = [None]
        ws = _make_sku_sheet(row1, row2)
        col_map = _build_col_index(ws)
        assert "bank_settlement_projected" not in col_map
        assert "net_units" not in col_map


# ── _parse_summary_sheet ──────────────────────────────────────────────────────

def _make_summary_sheet(rows: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    return wb2.active


class TestParseSummarySheet:

    def _standard_rows(self):
        return [
            ["Report Specifications", None, None],
            ["Orders Recieved During:", "2025-10-01 to 2025-10-31", None],
            ["PNL Summary", None, None],
            ["Item", "Amount (INR)", "Units"],
            ["Gross Sales", 324854, 813],
            ["Returns and Cancellations", -186090, -401],
            ["Estimated Net Sales", 138764, 412],
            ["Accounted Net Sales", 126397.5, None],
            ["Total Expenses", -37147.31, None],
            ["Bank Settlement (Projected)", 89480.52, None],
            ["Input Tax Credits", 6283.67, None],
            ["Earnings on Platform", 95764.19, None],
            ["Net Margin (% of Net Sales)", "75.76%", None],
            ["Already Paid (In Your Bank Account)", 89654.83, None],
            ["Pending (Flipkart to pay you)", -174.31, None],
        ]

    def test_gross_sales(self):
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        assert data["gross_sales"] == 324854.0
        assert data["gross_units"] == 813

    def test_returns(self):
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        assert data["returns_amount"] == -186090.0
        assert data["returned_units"] == 401  # stored as positive

    def test_net_sales_prefers_estimated(self):
        # estimated net sales comes before accounted — should be used
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        assert data["net_sales"] == 138764.0
        assert data["net_units"] == 412

    def test_bank_settlement(self):
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        assert data["bank_settlement"] == pytest.approx(89480.52, abs=0.01)

    def test_net_margin_string_pct(self):
        # "75.76%" string — must NOT multiply by 100
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        assert data["net_margin_pct"] == pytest.approx(75.76, abs=0.01)

    def test_net_margin_decimal(self):
        # 0.7576 decimal format (some Excel versions)
        rows = self._standard_rows()
        rows[12][1] = 0.7576  # replace "75.76%" with decimal
        ws = _make_summary_sheet(rows)
        data = _parse_summary_sheet(ws)
        assert data["net_margin_pct"] == pytest.approx(75.76, abs=0.01)

    def test_amount_settled_pending(self):
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        assert data["amount_settled"] == pytest.approx(89654.83, abs=0.01)
        assert data["amount_pending"] == pytest.approx(-174.31, abs=0.01)

    def test_all_fields_present(self):
        ws = _make_summary_sheet(self._standard_rows())
        data = _parse_summary_sheet(ws)
        required = ["gross_sales", "gross_units", "returns_amount", "returned_units",
                    "net_sales", "net_units", "total_expenses", "bank_settlement",
                    "input_tax_credits", "net_earnings", "net_margin_pct",
                    "amount_settled", "amount_pending"]
        for field in required:
            assert field in data, f"Missing field: {field}"


# ── extract_period_from_bytes ─────────────────────────────────────────────────

class TestExtractPeriod:

    def _make_file_with_period(self, period_str: str) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Report Type:", "Profit & Loss Report"])
        ws.append(["Orders Recieved During:", period_str])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_standard_period(self):
        data = self._make_file_with_period("2025-10-01 to 2025-10-31")
        from datetime import date
        start, end = extract_period_from_bytes(data)
        assert start == date(2025, 10, 1)
        assert end == date(2025, 10, 31)

    def test_different_month(self):
        data = self._make_file_with_period("2026-03-01 to 2026-03-31")
        from datetime import date
        start, end = extract_period_from_bytes(data)
        assert start == date(2026, 3, 1)
        assert end == date(2026, 3, 31)

    def test_missing_period_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Random", "Data"])
        buf = BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="Could not find the report period"):
            extract_period_from_bytes(buf.getvalue())
