"""
Integration test: parse the REAL Flipkart Excel file and assert exact values.
Skipped automatically if file not present (CI-safe).
"""
import pytest
from pathlib import Path
from io import BytesIO
import openpyxl

REAL_EXCEL = Path(r"C:\Users\MSI-PC\Downloads\Profit and Loss Report_04a22456-ee32-402b-839f-b28843fcb4ab.xlsx")

pytestmark = pytest.mark.skipif(
    not REAL_EXCEL.exists(),
    reason="Real Flipkart Excel not available"
)

from app.services.pnl import (
    _parse_summary_sheet, _build_col_index, _parse_sku_sheet,
    extract_period_from_bytes,
)
from datetime import date


@pytest.fixture(scope="module")
def file_bytes():
    return REAL_EXCEL.read_bytes()


@pytest.fixture(scope="module")
def workbook(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    return wb


@pytest.fixture(scope="module")
def summary(workbook):
    return _parse_summary_sheet(workbook[workbook.sheetnames[0]])


@pytest.fixture(scope="module")
def sku_rows(workbook):
    sku_ws = workbook[workbook.sheetnames[1]]
    col_map = _build_col_index(sku_ws)
    return _parse_sku_sheet(sku_ws, col_map), col_map


# ── Period extraction ─────────────────────────────────────────────────────────

def test_period_extraction(file_bytes):
    start, end = extract_period_from_bytes(file_bytes)
    assert start == date(2025, 10, 1)
    assert end == date(2025, 10, 31)


# ── Summary sheet — all 13 fields ────────────────────────────────────────────

def test_summary_gross_sales(summary):
    assert summary["gross_sales"] == pytest.approx(324854, abs=1)

def test_summary_gross_units(summary):
    assert summary["gross_units"] == 813

def test_summary_returns_amount(summary):
    assert summary["returns_amount"] == pytest.approx(-186090, abs=1)

def test_summary_returned_units(summary):
    assert summary["returned_units"] == 401  # stored as positive

def test_summary_net_sales(summary):
    assert summary["net_sales"] == pytest.approx(138764, abs=1)

def test_summary_net_units(summary):
    assert summary["net_units"] == 412

def test_summary_total_expenses(summary):
    assert summary["total_expenses"] == pytest.approx(-37147.31, abs=0.1)

def test_summary_bank_settlement(summary):
    assert summary["bank_settlement"] == pytest.approx(89480.52, abs=0.1)

def test_summary_input_tax_credits(summary):
    assert summary["input_tax_credits"] == pytest.approx(6283.67, abs=0.1)

def test_summary_net_earnings(summary):
    assert summary["net_earnings"] == pytest.approx(95764.19, abs=0.1)

def test_summary_net_margin_pct(summary):
    # Must be 75.76, NOT 7576 (safe_pct bug we fixed)
    assert summary["net_margin_pct"] == pytest.approx(75.76, abs=0.1)

def test_summary_amount_settled(summary):
    assert summary["amount_settled"] == pytest.approx(89654.83, abs=0.1)

def test_summary_amount_pending(summary):
    assert summary["amount_pending"] == pytest.approx(-174.31, abs=0.1)


# ── SKU sheet — column mapping ────────────────────────────────────────────────

def test_col_map_sku_id(sku_rows):
    _, col_map = sku_rows
    assert "sku_id" in col_map
    assert col_map["sku_id"] == 0

def test_col_map_cancelled_is_col_6_not_col_3(sku_rows):
    _, col_map = sku_rows
    # col 3 = "Returned & Cancelled Units" (total=40 for N65)
    # col 6 = "Cancellations" (actual=4 for N65) — MUST be 6
    assert col_map["cancelled_units"] == 6

def test_col_map_fixed_fee_present(sku_rows):
    _, col_map = sku_rows
    assert "fixed_fee" in col_map
    assert col_map["fixed_fee"] == 15

def test_col_map_bank_settlement_col_37(sku_rows):
    _, col_map = sku_rows
    # Must pick col 37 (first "[Projected]"), not col 45 (duplicate)
    assert col_map["bank_settlement_projected"] == 37


# ── SKU sheet — N65-WHITE row exact values ────────────────────────────────────

@pytest.fixture(scope="module")
def n65_row(sku_rows):
    rows, _ = sku_rows
    for r in rows:
        if r["platform_sku_name"] == "SHJ-JS-VRI-N65-WHITE":
            return r
    pytest.fail("N65-WHITE not found in parsed rows")


def test_n65_gross_units(n65_row):        assert n65_row["gross_units"] == 74
def test_n65_rto_units(n65_row):          assert n65_row["rto_units"] == 28
def test_n65_rvp_units(n65_row):          assert n65_row["rvp_units"] == 8
def test_n65_cancelled_units(n65_row):    assert n65_row["cancelled_units"] == 4   # was 40 before fix
def test_n65_net_units(n65_row):          assert n65_row["net_units"] == 34
def test_n65_fixed_fee(n65_row):          assert n65_row["fixed_fee"] == pytest.approx(-210, abs=0.1)
def test_n65_reverse_shipping(n65_row):   assert n65_row["reverse_shipping_fee"] == pytest.approx(-1356, abs=0.1)
def test_n65_taxes_gst(n65_row):          assert n65_row["taxes_gst"] == pytest.approx(-281.88, abs=0.1)
def test_n65_taxes_tcs(n65_row):          assert n65_row["taxes_tcs"] == pytest.approx(-25.16, abs=0.1)
def test_n65_taxes_tds(n65_row):          assert n65_row["taxes_tds"] == pytest.approx(-5.08, abs=0.1)
def test_n65_bsp(n65_row):                assert n65_row["bank_settlement_projected"] == pytest.approx(3320.98, abs=0.1)
def test_n65_input_tax_credits(n65_row):  assert n65_row["input_tax_credits"] == pytest.approx(312.12, abs=0.1)
def test_n65_net_earnings(n65_row):       assert n65_row["net_earnings"] == pytest.approx(3633.1, abs=0.1)
def test_n65_earnings_per_unit(n65_row):  assert n65_row["earnings_per_unit"] == pytest.approx(106.86, abs=0.1)

def test_n65_net_margin_pct_not_multiplied(n65_row):
    # Was showing 6988% before fix — must be 69.88
    assert n65_row["net_margin_pct"] == pytest.approx(69.88, abs=0.1)
    assert n65_row["net_margin_pct"] < 200  # sanity: never > 200%


# ── Totals cross-check: SKU rows sum must equal summary ──────────────────────

def test_sku_totals_gross_units_match_summary(sku_rows, summary):
    rows, _ = sku_rows
    total = sum(r["gross_units"] or 0 for r in rows)
    assert total == summary["gross_units"]

def test_sku_totals_net_units_match_summary(sku_rows, summary):
    rows, _ = sku_rows
    total = sum(r["net_units"] or 0 for r in rows)
    assert total == summary["net_units"]

def test_sku_totals_bsp_match_summary(sku_rows, summary):
    rows, _ = sku_rows
    total = sum(r["bank_settlement_projected"] or 0 for r in rows)
    assert total == pytest.approx(summary["bank_settlement"], abs=1.0)

def test_sku_totals_net_earnings_match_summary(sku_rows, summary):
    rows, _ = sku_rows
    total = sum(r["net_earnings"] or 0 for r in rows)
    assert total == pytest.approx(summary["net_earnings"], abs=1.0)

def test_sku_count(sku_rows):
    rows, _ = sku_rows
    assert len(rows) == 107

def test_no_sku_missing_name(sku_rows):
    rows, _ = sku_rows
    for r in rows:
        assert r["platform_sku_name"] and r["platform_sku_name"].strip()
